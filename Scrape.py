import requests
from bs4 import BeautifulSoup
import pprint
import csv
import openpyxl
from openpyxl import load_workbook
import BigScrape

'''writes header to xlsx file'''

#with open ('ScrapeFile.csv', 'w', encoding='utf-8') as csv_file:
#    csv_file.writerow(['Name', 'Smile', 'Molar Weight', 'Van-der-Waals', 'Molar Volume'])
wb = openpyxl.load_workbook('ScrapeFile.xlsx')
ws = wb.active
ws.cell(column=1, row=ws.max_row, value='name')
ws.cell(column=2, row=ws.max_row, value='smile')
ws.cell(column=3, row=ws.max_row, value='MolarWeight')
ws.cell(column=4, row=ws.max_row, value='MolarVol')
ws.cell(column=5, row=ws.max_row, value='Index of Refraction')
ws.cell(column=6, row=ws.max_row, value='Link')
wb.save('ScrapeFile.xlsx')

'''writes all polymer info to excel csv file'''

def my_scrape(URL):

    page = requests.get(URL)
    pp = pprint.PrettyPrinter(indent = 4)
    soup = BeautifulSoup(page.content, 'html.parser')

    results = soup.find(id = 'container')
    property_elems = results.find_all('section', class_ = 'datagrid')
    #for property_elem in property_elems:
        #print(property_elem, end='\n'*2)

    for results in soup.find_all('div',class_='datagrid'):
        smile = results.table.tbody
        #print(smile.text)
        #print()

    '''Getting the 4 paragraphs'''

    first = soup.find_all('div',class_='datagrid')[0]
    second = soup.find_all('div',class_='datagrid')[1]
    third = soup.find_all('div',class_='datagrid')[2]
    if 3 < len(soup.find_all('div',class_='datagrid')):
        fourth = soup.find_all('div',class_='datagrid')[3]
    else:
        fourth = soup.find_all('div',class_='datagrid')[2]


    fourthTR = fourth.find_all('tr')


    name = first.find_all('td')[3].text
    #print('Name:', name)

    smile = second.find_all('td')[5].text
    #print('Smile:', smile)

    '''getting molar weight from 4th paragraph'''

    MolarWeight = fourthTR[1]
    MolarWeightTD = MolarWeight.find_all('td')
    if MolarWeightTD[len(MolarWeightTD) - 1].text == '':
        MolarWeight = MolarWeightTD[len(MolarWeightTD) - 2].text
    else:
        MolarWeight = MolarWeightTD[len(MolarWeightTD) - 1].text
    #print('Molar Weight:', MolarWeight)

    '''van-der-waals volume'''
    VanDerWaals = fourthTR[2]
    VanDerWaalsTD = VanDerWaals.find_all('td')
    if VanDerWaalsTD[len(VanDerWaalsTD) - 1].text == '':
        VanDerWaals = VanDerWaalsTD[len(VanDerWaalsTD) - 2].text
    else:
        VanDerWaals = VanDerWaalsTD[len(VanDerWaalsTD) - 1].text
    #print('Van-der-Waals Volume:', VanDerWaals)

    '''Molar Volume'''

    MolarVol = fourthTR[3]
    MolarVolTD = MolarVol.find_all('td')
    if MolarVolTD[len(MolarVolTD) - 1].text == '':
        MolarVol = MolarVolTD[len(MolarVolTD) - 2].text
    else:
        MolarVol = MolarVolTD[len(MolarVolTD) - 1].text

    '''Index of refraction'''

    IndexOfRefraction = fourthTR[10]
    IndexOfRefractionTD = IndexOfRefraction.find_all('td')
    if IndexOfRefractionTD[len(IndexOfRefractionTD) - 1].text == '':
        IndexOfRefraction = IndexOfRefractionTD[len(IndexOfRefractionTD) - 2].text
    else:
        IndexOfRefraction = IndexOfRefractionTD[len(IndexOfRefractionTD) - 1].text


    ws.cell(column=1, row=ws.max_row+1, value=name)
    ws.cell(column=2, row=ws.max_row, value=smile)
    ws.cell(column=3, row=ws.max_row, value=MolarWeight)
    ws.cell(column=4, row=ws.max_row, value=MolarVol)
    ws.cell(column=5, row=ws.max_row, value=IndexOfRefraction)
    ws.cell(column=6, row=ws.max_row, value=URL)

    wb.save('ScrapeFile.xlsx')

    #value_list = [name, smile, MolarWeight, VanDerWaals, MolarVol]
    #csv_writer.writerow(value_list)

    #csv_file.close()
for yeet in BigScrape.finalLinks:
    print(yeet)
for endLink in BigScrape.finalLinks:
    newRequest = requests.get(endLink)
    if newRequest.status_code == 200:
        my_scrape(endLink)
    else:
        continue
