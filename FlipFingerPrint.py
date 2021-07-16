import requests
import csv
import openpyxl
from openpyxl import load_workbook


'''opens excel file'''

wb = load_workbook('fliptest.xlsx')
ws = wb.active
for row in ws.iter_rows(min_row=2):

    '''if both smiles are present, creates a flipped line'''
    if row[3].value:
        if row[4].value:
                newID1 = row[0].value + '_01'
                newID2 = row[0].value + '_02'
                newRow1 = (newID1, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value)
                newRow2 = (newID2, row[2].value, row[4].value, row[3].value, row[6].value, row[5].value)

                ws.append(newRow1)
                ws.append(newRow2)

                '''if only the first smile is present, creates 50/50 line'''
        else:
            newID3 = row[0].value + '01'
            newRow3 = (newID3, row[2].value, row[3].value, row[3].value, 50, 50)
            ws.append(newRow3)

            '''if only the second smiles is present, creates a 50/50 line'''
    else:
        newID4 = row[0].value + '_01'
        newRow4 = (newID4, row[2].value, row[4].value, row[4].value, 50, 50)
        ws.append(newRow4)

wb.save('fliptest.xlsx')
